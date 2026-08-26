# -*- coding: utf-8 -*-
"""
Excel'дан тизимга маълумот импорти — «Тоифаси ўзгарган / ўзгармаган».

ИШЛАТИШ:
    python import_excel.py "C:\\path\\Тоифаси ўзгарган-ўзгармаган.xlsx"
    python import_excel.py "...xlsx" --apply     # файлларга ЁЗАДИ

--apply берилмаса — ҳеч нарса ёзилмайди, фақат солиштирма ҳисобот чиқади.

НИМА ҚИЛАДИ
  1. Иккала варақни ўқийди (сарлавҳа 4-қатор, маълумот 6-қатордан).
  2. ЖАМИ сатрларини ташлаб юборади — уларни участка деб қабул қилиш
     майдонни икки баробар кўрсатиб юборарди.
  3. Устун номларини кодга мослайди (2 та ном Excel'да фарқ қилади).
  4. Учта файлни БИРГА янгилайди, боғланишни сақлаган ҳолда:
       sections_data.js   — асосий қаторлар (hk / changed)
       changed_data_v4.js — CHANGED_RAW: йиллар индекси шундан қурилади
       changed_extra.js   — CHANGED_EXTRA: _row бўйича захира қийматлар
     Учаласи _row орқали боғланган. Фақат бирини янгилаш харитадаги
     пинлар билан рақамларни бир-бирига мос келмайдиган қилиб қўяди.

_row = Excel'даги ҲАҚИҚИЙ қатор рақами (манбага изма-из қайтиш учун),
id   = "бўлим-_row".
"""
import io, json, os, re, shutil, sys, collections
from openpyxl import load_workbook

try: sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception: pass

HDR_ROW, SUB_ROW, DATA_ROW = 4, 5, 6

# Excel'даги ном  ->  код кутадиган ном.
# Фақат ФАРҚ ҚИЛАДИГАНЛАРИ; қолгани ўз номи билан кетади.
RENAME = {
  'Доимий фойдаланишга  ажратилган бино-иншоот (га)':
      'Доимий фойдаланишга тўғридан-тўғри ажратилган бино-иншоот (га)',
  'Ер қаъри участкаларини тадбиркорларга корхоналарга тўғридан-тўғри келишув асосида Ажратилган (га)':
      'Ер қаъри участкаларини корхоналарга тўғридан-тўғри келишув асосида Ажратилган (га)',
  'Ер участкаси ажратилган ёки талабгор корхона (ташкилот) номи':
      'Ер участкаси ажратилган корхона номи',
}

def num(v):
    if v is None: return 0.0
    try: return float(str(v).replace(' ', '').replace(',', '.'))
    except Exception: return 0.0

def is_totals_row(ws, row, hdr4):
    """5-qator ost-sarlavhami yoki JAMI satrimi?

    Ikkita varaq ikki xil: "ozgarmagan"da 5-qator haqiqiy ost-sarlavha
    ("shundan" -> Sug'oriladigan/Lalmi/yaylovlar), "ozgargan"da esa jami
    satri (T/r=470, Hudud nomi="Jami", Gektar=84385...). Uni sarlavha deb
    qabul qilsak, ustun nomi "Hudud nomi / Jami" bo'lib buziladi.
    """
    for c, name in enumerate(hdr4, 1):
        if name and 'Ҳудуд номи' in name:
            v = str(ws.cell(row=row, column=c).value or '').strip().lower()
            if v in ('жами', 'жами:', 'total'): return True
    first = ws.cell(row=row, column=1).value
    return isinstance(first, (int, float))

def headers(ws):
    """Sarlavhani quradi; kerak bo'lsa 5-qator ost-sarlavhasini qo'shadi."""
    spread = {}
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= HDR_ROW <= rng.max_row:
            v = ws.cell(row=rng.min_row, column=rng.min_col).value
            for c in range(rng.min_col, rng.max_col + 1):
                spread[c] = v
    hdr4 = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HDR_ROW, column=c).value or spread.get(c)
        hdr4.append(str(v).replace(chr(10), ' ').strip() if v else '')
    use_sub = not is_totals_row(ws, SUB_ROW, hdr4)
    out = []
    for c in range(1, ws.max_column + 1):
        m = hdr4[c-1]
        s = ''
        if use_sub:
            sv = ws.cell(row=SUB_ROW, column=c).value
            s = str(sv).replace(chr(10), ' ').strip() if sv else ''
        name = (m + (' / ' + s if s and s != m else '')).strip()
        # Кетма-кет бўш жойларни биттага келтирамиз. Excel'да устун номлари
        # кўпинча икки бўш жой билан ёзилган ('Жами ер майдони  (га)'), код эса
        # биттаси билан қидиради — шу сабабли майдон умуман ҳисобланмай қоларди.
        name = re.sub(r'\s+', ' ', name)
        out.append(RENAME.get(name, name))
    return out

def read_sheet(ws, district_col):
    hdr = headers(ws)
    if district_col not in hdr:
        sys.exit('XATO: "%s" ustuni topilmadi. Mavjud: %s' % (district_col, hdr[:6]))
    rows, skipped = [], 0
    for r in range(DATA_ROW, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or str(v).strip() == '' for v in vals):
            continue
        rec = {}
        for i, name in enumerate(hdr):
            if not name: continue
            v = vals[i]
            if v is None or (isinstance(v, str) and not v.strip()): continue
            rec[name] = v.strip() if isinstance(v, str) else v
        # ЖАМИ сатри: туман номи йўқ ёки «Жами» — участка эмас
        d = str(rec.get(district_col, '')).strip()
        if not d or d.lower() in ('жами', 'жами:', 'итого', 'total'):
            skipped += 1
            continue
        rec['_row'] = r
        rows.append(rec)
    return rows, skipped, hdr

def jsonify(v):
    """Excel qiymatini JSON uchun tayyorlaydi (sana/vaqt -> matn)."""
    if v is None or isinstance(v, (str, int, float, bool)): return v
    return str(v)

def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    src = sys.argv[1]
    apply = '--apply' in sys.argv
    if not os.path.exists(src): sys.exit('XATO: fayl topilmadi: ' + src)

    wb = load_workbook(src, data_only=True)
    sh_un = next((n for n in wb.sheetnames if 'ўзгармаган' in n or 'узгармаган' in n), None)
    sh_ch = next((n for n in wb.sheetnames if n not in (sh_un,) and 'ўзгарган' in n), None)
    if not sh_un or not sh_ch:
        sys.exit('XATO: varaqlar topilmadi. Mavjud: %s' % wb.sheetnames)

    hk_rows,  hk_skip,  hk_hdr  = read_sheet(wb[sh_un], 'Ҳудуд номи')
    ch_rows,  ch_skip,  ch_hdr  = read_sheet(wb[sh_ch], 'Ҳудуд номи')
    wb.close()

    for r in hk_rows: r['id'] = 'hk-%d' % r['_row']
    for r in ch_rows: r['id'] = 'changed-%d' % r['_row']

    # ---------- solishtirma ----------
    cur = json.loads(io.open('sections_data.js', encoding='utf-8-sig').read()
                     .split('=', 1)[1].strip().rstrip(';'))
    old_hk, old_ch = cur['sections']['hk'], cur['sections']['changed']
    o_hk_ha = sum(num(r.get('Жами ер майдони (га)')) for r in old_hk)
    o_ch_ha = sum(num(r.get('Гектар')) for r in old_ch)
    n_hk_ha = sum(num(r.get('Жами ер майдони  (га)') or r.get('Жами ер майдони (га)')) for r in hk_rows)
    n_ch_ha = sum(num(r.get('Гектар')) for r in ch_rows)

    line = '%-24s %7s %11s | %7s %11s | %8s %12s'
    print(line % ('', 'YANGI', 'ga', 'HOZIR', 'ga', 'FARQ', 'ga'))
    print('-' * 92)
    f = '%-24s %7d %11.1f | %7d %11.1f | %+8d %+12.1f'
    print(f % ('Тоифаси ўзгармаган (hk)', len(hk_rows), n_hk_ha, len(old_hk), o_hk_ha,
               len(hk_rows) - len(old_hk), n_hk_ha - o_hk_ha))
    print(f % ('Тоифаси ўзгарган', len(ch_rows), n_ch_ha, len(old_ch), o_ch_ha,
               len(ch_rows) - len(old_ch), n_ch_ha - o_ch_ha))
    print(f % ('JAMI', len(hk_rows) + len(ch_rows), n_hk_ha + n_ch_ha,
               len(old_hk) + len(old_ch), o_hk_ha + o_ch_ha,
               len(hk_rows) + len(ch_rows) - len(old_hk) - len(old_ch),
               n_hk_ha + n_ch_ha - o_hk_ha - o_ch_ha))
    print()
    print('tashlangan JAMI satrlari: ўзгармаган %d, ўзгарган %d' % (hk_skip, ch_skip))
    print('ustunlar: ўзгармаган %d, ўзгарган %d' % (
        len([h for h in hk_hdr if h]), len([h for h in ch_hdr if h])))

    if not apply:
        print()
        print('>>> SINOV REJIMI — hech narsa yozilmadi. Yozish uchun: --apply')
        return

    # ---------- zaxira ----------
    for fn in ('sections_data.js', 'changed_data_v4.js', 'changed_extra.js'):
        shutil.copy2(fn, fn + '.bak')
    print('\nzaxira: *.js.bak')

    # ---------- 1. sections_data.js ----------
    cur['sections']['hk'] = [{k: jsonify(v) for k, v in r.items()} for r in hk_rows]
    cur['sections']['changed'] = [{k: jsonify(v) for k, v in r.items()} for r in ch_rows]
    io.open('sections_data.js', 'w', encoding='utf-8-sig', newline='').write(
        'window.SECTIONS_RAW_FULL = ' + json.dumps(cur, ensure_ascii=False, separators=(',', ':')) + ';')

    # ---------- 2. changed_data_v4.js (CHANGED_RAW) ----------
    # Faqat kod ishlatadigan maydonlar: id,t,m,q,k,kr,g,y,p,pr,tr,au,sg,ql,dm,kt,iz
    G = lambda r, k: ('' if r.get(k) is None else str(r.get(k)).strip())
    raw = []
    # changedRawRow() `id + 5 === _row` боғланишига таянади — акс ҳолда
    # CHANGED_RAW ёзуви бошқа участкага уланиб қолади.
    for r in ch_rows:
        raw.append({
            'id': r['_row'] - 5,
            't':  G(r, 'Ҳудуд номи'),
            'm':  G(r, 'МФЙ номи'),
            'q':  G(r, 'Фойдаланиш мақсади'),
            'k':  G(r, 'Ер участкаси ажратилган корхона номи'),
            # kr — КООРДИНАТА матни. Код уни makePlotCandidate'га узатади,
            # шу боис бу ерга 'Контур рақами' тушса харитада пин чиқмайди.
            'kr': G(r, 'Координата рақами'),
            'g':  G(r, 'Гектар'),
            'y':  G(r, 'Тоифаси ўзгарган йили (2019,2020, 2021,2022, 2023,2024,2025)'),
            'p':  G(r, 'Тоифаси ўзгарган Протокол ёки ВМ фармойишга асосан'),
            'pr': G(r, 'Протокол рақами'),
            'tr': G(r, 'Бино-иншоот/Кон'),
            'kt': G(r, 'Коннинг тури (Норуда, Сочма, Стратегик, НКМК, Уран)'),
            'au': G(r, 'Бино-иншоот аукцион савдоларига чиқарилган (га)'),
            'sg': G(r, 'Бино-иншоот аукцион савдоларида сотилган ер майдони (га)'),
            'ql': G(r, 'Бино-иншоот қолдиқ (га)'),
            'dm': G(r, 'Доимий фойдаланишга тўғридан-тўғри ажратилган бино-иншоот (га)'),
            'iz': G(r, 'Изоҳ'),
        })
    io.open('changed_data_v4.js', 'w', encoding='utf-8-sig', newline='').write(
        'const CHANGED_RAW=' + json.dumps(raw, ensure_ascii=False, separators=(',', ':')) + ';')

    # ---------- 3. changed_extra.js (CHANGED_EXTRA) ----------
    # _row bo'yicha zaxira qiymatlar. Yangi _row raqamlariga MOSLANADI —
    # aks holda eski kalitlar boshqa uchastkaga tegib, noto'g'ri ko'rsatardi.
    extra = {}
    for r in ch_rows:
        extra[str(r['_row'])] = {
            'm':  G(r, 'МФЙ номи'),
            'kt': G(r, 'Коннинг тури (Норуда, Сочма, Стратегик, НКМК, Уран)'),
            'bd': num(r.get('Доимий фойдаланишга тўғридан-тўғри ажратилган бино-иншоот (га)')),
            'kd': num(r.get('Ер қаъри участкаларини корхоналарга тўғридан-тўғри келишув асосида Ажратилган (га)')),
            'kn': num(r.get('Ер қаърига доимий фойдаланишга ажратилмаган (га)')),
            'kp': num(r.get('Ер қаъри аукционга  чиқарилган (га)') or r.get('Ер қаъри аукционга чиқарилган (га)')),
            'ks': num(r.get('Ер қаъри аукционда сотилган (га)')),
        }
    io.open('changed_extra.js', 'w', encoding='utf-8-sig', newline='').write(
        'window.CHANGED_EXTRA=' + json.dumps(extra, ensure_ascii=False, separators=(',', ':')) + ';')

    print('yozildi: sections_data.js, changed_data_v4.js, changed_extra.js')
    print('CHANGED_RAW: %d yozuv | CHANGED_EXTRA: %d kalit' % (len(raw), len(extra)))

if __name__ == '__main__':
    main()
